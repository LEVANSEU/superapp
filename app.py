import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
import re

st.set_page_config(layout="wide")
st.title("Excel გენერატორი")

# გაფართოების CSS ჰაკი
st.markdown("""
    <style>
        .main { max-width: 95%; padding-left: 2rem; padding-right: 2rem; }
        .block-container { padding-top: 1rem; padding-bottom: 1rem; }
        button[kind="secondary"] { width: 100%; }
    </style>
""", unsafe_allow_html=True)

report_file = st.file_uploader("ატვირთე ანგარიშფაქტურების ფაილი (report.xlsx)", type=["xlsx"])
statement_file = st.file_uploader("ატვირთე საბანკო ამონაწერის ფაილი (statement.xlsx)", type=["xlsx"])

if report_file and statement_file:
    purchases_df = pd.read_excel(report_file, sheet_name='Grid')
    bank_df = pd.read_excel(statement_file)

    purchases_df['დასახელება'] = purchases_df['გამყიდველი'].astype(str).apply(lambda x: re.sub(r'^\(\d+\)\s*', '', x).strip())
    purchases_df['საიდენტიფიკაციო კოდი'] = purchases_df['გამყიდველი'].apply(lambda x: ''.join(re.findall(r'\d', str(x)))[:11])
    bank_df['P'] = bank_df.iloc[:, 15].astype(str).str.strip()
    bank_df['Amount'] = pd.to_numeric(bank_df.iloc[:, 3], errors='coerce').fillna(0)

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet(title="ანგარიშფაქტურები კომპანიით")
    ws1.append(['დასახელება', 'საიდენტიფიკაციო კოდი', 'ანგარიშფაქტურის №', 'ანგარიშფაქტურის თანხა', 'ჩარიცხული თანხა'])

    company_summaries = []

    for company_id, group in purchases_df.groupby('საიდენტიფიკაციო კოდი'):
        company_name = group['დასახელება'].iloc[0]
        unique_invoices = group.groupby('სერია №')['ღირებულება დღგ და აქციზის ჩათვლით'].sum().reset_index()
        company_invoice_sum = unique_invoices['ღირებულება დღგ და აქციზის ჩათვლით'].sum()

        company_summary_row = ws1.max_row + 1
        payment_formula = f"=SUMIF(საბანკოამონაწერი!P:P, B{company_summary_row}, საბანკოამონაწერი!D:D)"
        ws1.append([company_name, company_id, '', company_invoice_sum, payment_formula])

        for _, row in unique_invoices.iterrows():
            ws1.append(['', '', row['სერია №'], row['ღირებულება დღგ და აქციზის ჩათვლით'], ''])

        company_summaries.append((company_name, company_id, company_invoice_sum))

    for sheet_title, content_df in [
        ("დეტალური მონაცემები", purchases_df),
        ("საბანკოამონაწერი", bank_df),
        ("ანგარიშფაქტურის დეტალები", purchases_df[['სერია №', 'საქონელი / მომსახურება', 'ზომის ერთეული', 'რაოდ.', 'ღირებულება დღგ და აქციზის ჩათვლით']].rename(columns={'სერია №': 'ინვოისის №'})),
        ("გადარიცხვები_უბმოლოდ", bank_df[~bank_df['P'].isin(purchases_df['საიდენტიფიკაციო კოდი'])]),
        ("განახლებული ამონაწერი", bank_df),
    ]:
        ws = wb.create_sheet(title=sheet_title)
        ws.append(content_df.columns.tolist())
        for row in content_df.itertuples(index=False):
            ws.append(row)

    ws7 = wb.create_sheet(title="კომპანიების_ჯამები")
    ws7.append(['დასახელება', 'საიდენტიფიკაციო კოდი', 'ანგარიშფაქტურების ჯამი', 'ჩარიცხული თანხა'])
    for idx, (company_name, company_id, invoice_sum) in enumerate(company_summaries, start=2):
        payment_formula = f"=SUMIF(საბანკოამონაწერი!P:P, B{idx}, საბანკოამონაწერი!D:D)"
        ws7.append([company_name, company_id, invoice_sum, payment_formula])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # მთავარი ხედის ან დეტალების ჩვენება
    if 'selected_company' not in st.session_state:
        st.subheader("📋 კომპანიების ჩამონათვალი")

        for name, company_id, invoice_sum in company_summaries:
            col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 1.5])
            with col1:
                st.markdown(name)
            with col2:
                if st.button(f"{company_id}", key=f"id_{company_id}"):
                    st.session_state['selected_company'] = company_id

            paid_sum = bank_df[bank_df["P"] == str(company_id)]["Amount"].sum()
            difference = invoice_sum - paid_sum

            with col3:
                st.write(f"{invoice_sum:,.2f}")
            with col4:
                st.write(f"{paid_sum:,.2f}")
            with col5:
                st.write(f"{difference:,.2f}")

    else:
        selected_code = st.session_state['selected_company']
        df_full = pd.read_excel(report_file, sheet_name='Grid')
        df_full['დასახელება'] = df_full['გამყიდველი'].astype(str).apply(lambda x: re.sub(r'^\(\d+\)\s*', '', x).strip())
        df_full['საიდენტიფიკაციო კოდი'] = df_full['გამყიდველი'].apply(lambda x: ''.join(re.findall(r'\d', str(x)))[:11])
        matching_df = df_full[df_full['საიდენტიფიკაციო კოდი'] == selected_code]

        if not matching_df.empty:
            company_name = matching_df['დასახელება'].iloc[0]
            st.subheader(f"🔎 ({selected_code}) {company_name} - ანგარიშფაქტურები")
            st.dataframe(matching_df, use_container_width=True)

            # საძიებო ველი
            st.subheader("🔍 მოძებნე გუგლში მასალა ან მომსახურება")
            col1, col2 = st.columns([3, 1])
            with col1:
                search_term = st.text_input("ჩაწერე სახელი ან სიტყვა:")
            with col2:
                if st.button("ძებნა"):
                    if search_term.strip():
                        search_url = f"https://www.google.com/search?q={search_term.replace(' ', '+')}"
                        st.markdown(f"[🌐 გადადი გუგლზე]({search_url})", unsafe_allow_html=True)
                    else:
                        st.warning("გთხოვ ჩაწერე ტექსტი ძებნამდე.")

            # Excel ფაილის ჩამოტვირთვა
            company_output = io.BytesIO()
            company_wb = Workbook()
            ws = company_wb.active
            ws.title = company_name[:31]
            ws.append(matching_df.columns.tolist())
            for row in matching_df.itertuples(index=False):
                ws.append(row)
            company_wb.save(company_output)
            company_output.seek(0)

            st.download_button(
                label=f"⬇️ ჩამოტვირთე {company_name} ინვოისების Excel",
                data=company_output,
                file_name=f"{company_name}_ინვოისები.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("📭 ჩანაწერი ვერ მოიძებნა ამ კომპანიაზე.")

        if st.button("⬅️ დაბრუნება სრულ სიაზე"):
            del st.session_state['selected_company']

    st.success("✅ ფაილი მზადაა! ჩამოტვირთე აქედან:")
    st.download_button(
        label="⬇️ ჩამოტვირთე Excel ფაილი",
        data=output,
        file_name="საბოლოო_ფაილი.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
